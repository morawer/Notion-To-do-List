from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, PatternFill, Font
import time
import tkinter as tk 
from tkinter import filedialog
#import msvcrt

def fileSelection():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename()
    return file_path

def acum_lines():
    
    try:
        excel = fileSelection()
        wb = load_workbook(excel)
        ws = wb.active


    except:
        print("ERROR - PROCESS CANCELED")
        quit()
    
    
    acum = 0
    acum_total = 0
    row_number = 2
    
    first = True
    
    while acum != 0 or first == True:
        
        if first == False:
            wb = load_workbook(
                "/home/dani/Projects/To_do_List_Excel/NUEVO-PANELES.xlsx")
            ws = wb.active
            acum = 0
    
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        
            co = ws.cell(row=row_number, column=1).value
            next_co = ws.cell(row=row_number+1, column=1).value
        
            line = ws.cell(row=row_number, column=2).value
            next_line = ws.cell(row=row_number+1, column=2).value

            item = ws.cell(row=row_number, column=3).value
            next_item = ws.cell(row=row_number+1, column=3).value

            qty = ws.cell(row=row_number, column=5).value
            next_qty = ws.cell(row=row_number+1, column=5).value
                    
            print(row_number)
        
            if (co == next_co) and (line == next_line) and (item == next_item):
                try:
                    ws.cell(row=row_number, column=5,
                           value=qty + next_qty)
                    ws.delete_rows(idx=row_number+1)
                    acum = acum+1
                    print('HOLA')
                except:
                    print("ERROR")
            
            row_number = row_number + 1
            
        wb.save(f"NUEVO-PANELES.xlsx")
        first = False
        acum_total = acum_total + acum
        print("ACUM= ", acum)
        row_number = 2
        
    print(acum_total)
            

starting_point = time.time()

acum_lines()
    

# grab the active worksheet
wb = load_workbook(
    "/home/dani/Projects/To_do_List_Excel/NUEVO-PANELES.xlsx")
ws = wb.active
co = 0
line = 0
qty = 0
row_counter = 0
id_line = 0
row_line = 2

title = True
last_value = 0
new_value = 0


for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    
    item_value = str(row[2].value)
    name_value = row[3].value
    qty_value = row[4].value

    co_value = row[0].value
    line_value = row[1].value
    
    next_co = ws.cell(row=row_line + 1, column=1).value
    next_line = ws.cell(row=row_line + 1, column=2).value
    
    new_value = item_value
    
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    my_fill = PatternFill(start_color="d3d3d3",
                          end_color="d3d3d3", fill_type="solid")

    cell_alignment = Alignment(horizontal="center", vertical="center")
    
    if (co_value != co or line_value != line):

        print(f"{co_value}-{line_value}")

        co = row[0].value
        line = row[1].value
        qty = qty + 1

        wb_panels = load_workbook(
            "/home/dani/Projects/To_do_List_Excel/PANELS/Plantilla Paneles.xlsx")

        ws_panels = wb_panels.active
        ws_panels["B2"].value = f"{co}-{line}"
        
        row_counter = 0
        
    if (title and last_value == 0) or (last_value[0:3] != new_value[0:3]) or (new_value.__contains__("S") and not last_value.__contains__("S")) or (last_value.__contains__("S") and not new_value.__contains__("S")):
        #TODO: Make titles for each new block of items
             
        ws_panels.merge_cells(start_row=11 + row_counter,
                              start_column=2, end_row=11 + row_counter, end_column=12)
        
        if new_value[0:3] == "500" and not item_value.__contains__("S"):
            ws_panels.cell(row=11 + row_counter, column=2,
                           value="PANELES GENIOX").alignment = cell_alignment
            ws_panels.cell(row=11 + row_counter, column=2).fill = my_fill

            title = False
            last_value = new_value
        
        elif new_value[0:3] == "503":
            ws_panels.cell(row=11 + row_counter, column=2,
                           value="PUERTAS GENIOX").alignment = cell_alignment
            ws_panels.cell(row=11 + row_counter, column=2).fill = my_fill

            title = False
            last_value = new_value

        
        elif new_value[0:3] == "506":
            ws_panels.cell(row=11 + row_counter, column=2,
                           value="PANELES PISO INTERMEDIO").alignment = cell_alignment
            ws_panels.cell(row=11 + row_counter, column=2).fill = my_fill

            title = False
            last_value = new_value

            
        elif new_value.__contains__("S"):
            ws_panels.cell(row=11 + row_counter, column=2,
                           value="PANELES GX ON").alignment = cell_alignment
            ws_panels.cell(row=11 + row_counter, column=2).fill = my_fill

            title = False
            last_value = new_value

            
        else:
            ws_panels.cell(row=11 + row_counter, column=2,
                           value="PANELES ESPECIALES").alignment = cell_alignment
            ws_panels.cell(row=11 + row_counter, column=2).fill = my_fill

            title = False
            last_value = new_value
            
        row_counter = row_counter + 1
         
    ws_panels.merge_cells(start_row=11 + row_counter,
                          start_column=2, end_row=11 + row_counter, end_column=3)
    
    ws_panels.merge_cells(start_row=11 + row_counter,
                          start_column=4, end_row=11 + row_counter, end_column=9)

    ws_panels.cell(row=11 + row_counter, column=2, 
                   value=item_value).alignment=cell_alignment
    ws_panels.cell(row=11 + row_counter, column=4,
                   value=name_value).alignment = cell_alignment
    ws_panels.cell(row=11 + row_counter, column=10,
                   value=qty_value).alignment = cell_alignment

    ws_panels.cell(row=11 + row_counter, column=2).border = thin_border
    ws_panels.cell(row=11 + row_counter, column=3).border = thin_border
    ws_panels.cell(row=11 + row_counter, column=4).border = thin_border
    ws_panels.cell(row=11 + row_counter, column=5).border = thin_border
    ws_panels.cell(row=11 + row_counter, column=6).border = thin_border
    ws_panels.cell(row=11 + row_counter, column=7).border = thin_border
    ws_panels.cell(row=11 + row_counter, column=8).border = thin_border
    ws_panels.cell(row=11 + row_counter, column=9).border = thin_border
    ws_panels.cell(row=11 + row_counter, column=10).border = thin_border
    ws_panels.cell(row=11 + row_counter, column=11).border = thin_border
    ws_panels.cell(row=11 + row_counter, column=12).border = thin_border
        
    row_counter = row_counter + 1
    
    print(f"{co} == {line} || {next_co} == {next_line}")

    if (co_value != next_co or line_value != line):
        
        try:
            print(f"{co_value}-{line_value}-PANELES.xlsx")
            wb_panels.save(f"{co_value}-{line_value}-PANELES.xlsx")
        
            new_value = 0

        except:
            print(f"{co_value}-{line_value} --- ERROR")
    
    row_line= row_line + 1
    
    
elapsed_time = time.time() - starting_point
elapsed_time_int = int(elapsed_time)
elapsed_time_minutes = elapsed_time_int / 60 
elapsed_time_seconds = elapsed_time_int % 60

print(f"TOTAL = {qty}")
print(f"Done in {elapsed_time_minutes:.0f} minutes and {elapsed_time_seconds} seconds.")
print()
print("Press any button to exit")
#msvcrt.getch()