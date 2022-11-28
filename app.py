from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, PatternFill, Font
import time
import tkinter as tk 
from tkinter import filedialog

def fileSelection():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename()
    return file_path

excel = fileSelection()

wb = load_workbook(excel)

# grab the active worksheet
ws = wb.active

co = 0
line = 0
qty = 0
row_counter = 0
id_line = 0


panels_array = []

starting_point = time.time()

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    
    item_value = str(row[2].value)
    name_value = row[3].value
    qty_value = row[4].value

    co_value = row[0].value
    line_value = row[1].value
    
    if (co_value != co or line_value != line):

        print(f"{co_value}-{line_value}")

        co = row[0].value
        line = row[1].value
        qty = qty + 1

        wb_panels = load_workbook("Plantilla Paneles.xlsx")

        ws_panels = wb_panels.active
        ws_panels["B2"].value = f"{co}-{line}"
        row_counter = 0
        
    ws_panels.merge_cells(start_row=11 + row_counter,
                          start_column=2, end_row=11 + row_counter, end_column=3)
    
    ws_panels.merge_cells(start_row=11 + row_counter,
                          start_column=4, end_row=11 + row_counter, end_column=9)

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    my_fill = PatternFill(start_color="d3d3d3", end_color="d3d3d3", fill_type="solid")
    
    cell_alignment = Alignment(horizontal="center", vertical="center")
    
    

    ws_panels.cell(row=11 + row_counter, column=2, 
                   value=item_value).alignment=cell_alignment
    ws_panels.cell(row=11 + row_counter, column=4,
                   value=name_value).alignment = cell_alignment
    ws_panels.cell(row=11 + row_counter, column=10,
                   value=qty_value).alignment = cell_alignment

    ws_panels.cell(row=11 + row_counter, column=2).border = thin_border
    ws_panels.cell(row=11 + row_counter, column=3).border = thin_border
    ws_panels.cell(row=11 + row_counter, column=4).border = thin_border
    ws_panels.cell(row=11 + row_counter, column=5).border = thin_border
    ws_panels.cell(row=11 + row_counter, column=6).border = thin_border
    ws_panels.cell(row=11 + row_counter, column=7).border = thin_border
    ws_panels.cell(row=11 + row_counter, column=8).border = thin_border
    ws_panels.cell(row=11 + row_counter, column=9).border = thin_border
    ws_panels.cell(row=11 + row_counter, column=10).border = thin_border
    ws_panels.cell(row=11 + row_counter, column=11).border = thin_border
    ws_panels.cell(row=11 + row_counter, column=12).border = thin_border

    if item_value.__contains__("S"):
        id_line = 1
        
    elif item_value[0:3] == "503":
        id_line = 2
       
    elif item_value[0:3] == "506":
        id_line = 3

    elif item_value[0:2] != "50":
        id_line = 5

    elif item_value[0:3] == "500" and not item_value.__contains__("S"):
        id_line = 6

    if (id_line%2 != 0):
        ws_panels.cell(row=11 + row_counter, column=2).fill= my_fill
        ws_panels.cell(row=11 + row_counter, column=3).fill= my_fill
        ws_panels.cell(row=11 + row_counter, column=4).fill= my_fill
        ws_panels.cell(row=11 + row_counter, column=5).fill= my_fill
        ws_panels.cell(row=11 + row_counter, column=6).fill= my_fill
        ws_panels.cell(row=11 + row_counter, column=7).fill= my_fill
        ws_panels.cell(row=11 + row_counter, column=8).fill= my_fill
        ws_panels.cell(row=11 + row_counter, column=9).fill= my_fill
        ws_panels.cell(row=11 + row_counter, column=10).fill= my_fill
        ws_panels.cell(row=11 + row_counter, column=11).fill= my_fill
        ws_panels.cell(row=11 + row_counter, column=12).fill= my_fill
        
    row_counter = row_counter + 1

    try:
        wb_panels.save(f"{co_value}-{line_value}-PANELES.xlsx")

    except:
        print(f"{co_value}-{line_value} --- ERROR")
    
    
elapsed_time = time.time() - starting_point
elapsed_time_int = int(elapsed_time)
elapsed_time_minutes = elapsed_time_int / 60 
elapsed_time_seconds = elapsed_time_int % 60

print(f"TOTAL = {qty}")
print(f"Done in {elapsed_time_minutes:.0f} minutes and {elapsed_time_seconds} seconds.")